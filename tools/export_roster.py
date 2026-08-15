#!/usr/bin/env python3
"""武将・必殺技・固有特性の一覧を xlsx と csv へ書き出す。

usage: python3 tools/export_roster.py [出力先ディレクトリ]

見るための出力であって、これを読み込んで動くものは無い。
元データは sim/roster.py の ROSTER と sim/cards.json（生成後の能力値）。
"""
import csv
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "sim"))

import roster  # noqa: E402
from engine import CARDS  # noqa: E402

TROOP = {"inf": "歩兵", "cav": "騎兵", "arc": "弓兵"}
ROLE = {"tank": "耐久", "bruiser": "均衡", "dps": "火力",
        "support": "支援", "burst": "瞬発"}
FACTION = {"wei": "魏", "shu": "蜀", "go": "呉", "gun": "群雄", "none": "中立"}
TIER = {"low": "低", "mid": "中", "high": "高"}

TARGET = {
    "front_enemy": "敵1体（最前）", "enemy_lowest": "敵1体（残兵力が最少）",
    "lane_enemies": "敵1列", "enemy_back_lane": "敵後列", "all_enemies": "敵全体",
    "lane_allies": "味方1列", "all_allies": "味方全体",
    "lowest_ally": "味方1体（残兵力が最少）", "self": "自分",
}
EFFECT = {
    "damage": "ダメージ", "dot": "継続ダメージ", "mod": "能力補正",
    "stun": "行動阻害", "heal": "回復", "gauge": "ゲージ付与",
}
STAT = {"atk": "攻撃力", "dfn": "防御力", "acc": "命中率", "speed": "移動速度"}

TRAIT_NOTE = {
    "vanguard": ("常在", "陣頭。前衛に置くと兵力が増える。総大将を前へ出せる唯一の札"),
    "vs_wei": ("常在", "対魏。魏の武将へのダメージ増（群雄にも当たる）"),
    "vs_shu": ("常在", "対蜀。蜀の武将へのダメージ増（群雄にも当たる）"),
    "vs_go": ("常在", "対呉。呉の武将へのダメージ増（群雄にも当たる）"),
}


def trait_rows():
    """固有特性の一覧。誘発型は TRIGGERS、常在型は engine 側の実装。"""
    used = {}
    for e in roster.ROSTER:
        for t in e[6]:
            used.setdefault(t, []).append(f"{e[0]}〔{e[1]}〕")
    rows = []
    for key in sorted(used, key=lambda k: (-len(used[k]), k)):
        if key in roster.TRIGGERS:
            t = roster.TRIGGERS[key]
            kind = "誘発"
            note = (f"{t['trigger']} で発動 / 対象 {TARGET.get(t['target'], t['target'])}"
                    f" / 1戦{t.get('limit', '-')}回まで")
            eff = " + ".join(describe_effect(f) for f in t["effects"])
            name = t["name"]
        else:
            kind, note = TRAIT_NOTE.get(key, ("常在", ""))
            eff, name = "", key
        rows.append([key, name, kind, eff, note, len(used[key]),
                     "、".join(used[key])])
    return rows


def describe_effect(f):
    kind = EFFECT.get(f["type"], f["type"])
    if f["type"] in ("damage", "dot"):
        s = f"{kind} 威力{f.get('power', '-')}%"
        if f["type"] == "dot":
            s += f"（{f.get('duration', 0) // 10}秒）"
        return s
    if f["type"] == "mod":
        sign = "+" if f.get("value", 0) >= 0 else ""
        return (f"{STAT.get(f.get('stat'), f.get('stat'))} {sign}{f.get('value')}%"
                f"（{f.get('duration', 0) // 10}秒）")
    if f["type"] == "stun":
        return f"{kind} {f.get('duration', 0) // 10}秒"
    if f["type"] == "heal":
        return f"{kind} 攻撃力の{f.get('power', '-')}%"
    if f["type"] == "gauge":
        return f"{kind} 自然増加の{f.get('seconds', '-')}秒ぶん"
    return kind


def skill_rows():
    rows = []
    for e in roster.ROSTER:
        s = e[5]
        rows.append([
            s["name"], f"{e[0]}〔{e[1]}〕", e[2],
            TARGET.get(s["target"], s["target"]),
            " + ".join(describe_effect(f) for f in s["effects"]),
            s.get("gauge", 100),
            len(s["effects"]),
        ])
    return rows


def card_rows():
    rows = []
    for e in roster.ROSTER:
        cid = f"{roster.ROMAJI[e[0]]}_{e[2]}"
        c = CARDS[cid]
        over = e[7] if len(e) > 7 else {}
        rows.append([
            c["name"], e[0], e[1], FACTION.get(c["faction"], c["faction"]),
            c["cost"], TIER.get(c["tier"], c["tier"]),
            TROOP.get(c["troop"], c["troop"]), ROLE.get(c["role"], c["role"]),
            c["hp"], c["atk"], c["dfn"], c["acc"], c["crit"], c["power"],
            e[5]["name"], e[5].get("gauge", 100),
            over.get("gauge_rate", c["gauge_rate"]), over.get("gauge_start", 0),
            "、".join(e[6]),
        ])
    return rows


SHEETS = [
    ("武将", ["名前", "人物", "字号", "勢力", "コスト", "帯", "兵種", "役割",
             "兵力", "攻撃力", "防御力", "命中率", "クリ率", "実力比%",
             "必殺技", "消費ゲージ%", "ゲージ上昇率", "初期ゲージ", "固有特性"],
     card_rows),
    ("必殺技", ["技名", "武将", "コスト", "対象", "効果", "消費ゲージ%", "効果数"],
     skill_rows),
    ("固有特性", ["キー", "名前", "型", "効果", "備考", "採用枚数", "持つ武将"],
     trait_rows),
]


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else os.getcwd()
    os.makedirs(out, exist_ok=True)
    made = []
    for name, head, fn in SHEETS:
        path = os.path.join(out, f"三国志_{name}一覧.csv")
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(head)
            w.writerows(fn())
        made.append(path)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("openpyxl が無いので csv だけ書いた")
        for p in made:
            print(" ", p)
        return
    wb = Workbook()
    wb.remove(wb.active)
    for name, head, fn in SHEETS:
        ws = wb.create_sheet(name)
        ws.append(head)
        for r in fn():
            ws.append(r)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4D78")
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        for i, _h in enumerate(head, 1):
            col = ws.cell(row=1, column=i).column_letter
            width = max(len(str(ws.cell(row=r, column=i).value or ""))
                        for r in range(1, ws.max_row + 1))
            ws.column_dimensions[col].width = min(max(width * 1.6, 8), 60)
        ws.auto_filter.ref = ws.dimensions
    path = os.path.join(out, "三国志_武将データ.xlsx")
    wb.save(path)
    made.append(path)
    for p in made:
        print(" ", p)


if __name__ == "__main__":
    main()
